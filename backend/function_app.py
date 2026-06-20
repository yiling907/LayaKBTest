import json
import logging
import os
import uuid
import io

import azure.functions as func

from shared import blob_client, search_client, openai_client, cosmos_client

app = func.FunctionApp(http_auth_level=func.AuthLevel.FUNCTION)

logger = logging.getLogger(__name__)

SYSTEM_PROMPT = """You are a helpful assistant that answers questions based solely on the provided context.
If the answer is not found in the context, say "I don't have enough information to answer that."
Always cite the source document(s) you used."""


# ---------------------------------------------------------------------------
# Health check
# ---------------------------------------------------------------------------

@app.route(route="health", methods=["GET"], auth_level=func.AuthLevel.ANONYMOUS)
def health(req: func.HttpRequest) -> func.HttpResponse:
    return func.HttpResponse(
        json.dumps({"status": "ok"}),
        mimetype="application/json",
        status_code=200,
    )


# ---------------------------------------------------------------------------
# Ingest a document
# ---------------------------------------------------------------------------

@app.route(route="ingest", methods=["POST"])
def ingest(req: func.HttpRequest) -> func.HttpResponse:
    file = req.files.get("file")
    if not file:
        return func.HttpResponse(
            json.dumps({"error": "No file provided. Send a multipart/form-data request with field 'file'."}),
            mimetype="application/json",
            status_code=400,
        )

    doc_id = str(uuid.uuid4())
    file_name = file.filename or f"{doc_id}.bin"
    raw_bytes = file.read()

    # 1. Store raw document in Blob Storage
    blob_client.upload_document(file_name, raw_bytes)

    # 2. Extract text
    text = _extract_text(file_name, raw_bytes)

    # 3. Chunk the text
    chunks = _chunk_text(text)

    # 4. Ensure search index exists
    search_client.ensure_index()

    # 5. Embed + index each chunk
    search_docs = []
    for i, chunk in enumerate(chunks):
        vector = openai_client.get_embedding(chunk)
        search_docs.append({
            "id": f"{doc_id}_{i}",
            "document_id": doc_id,
            "document_name": file_name,
            "content": chunk,
            "content_vector": vector,
        })
    search_client.upsert_chunks(search_docs)

    # 6. Save metadata to Cosmos DB
    metadata = {
        "id": doc_id,
        "name": file_name,
        "size": len(raw_bytes),
        "chunks": len(chunks),
        "status": "indexed",
    }
    cosmos_client.save_document_metadata(metadata)

    logger.info("Ingested document %s (%d chunks)", file_name, len(chunks))
    return func.HttpResponse(
        json.dumps(metadata),
        mimetype="application/json",
        status_code=200,
    )


# ---------------------------------------------------------------------------
# Query the knowledge base
# ---------------------------------------------------------------------------

@app.route(route="query", methods=["POST"])
def query(req: func.HttpRequest) -> func.HttpResponse:
    body = req.get_json()
    question = (body or {}).get("question", "").strip()
    if not question:
        return func.HttpResponse(
            json.dumps({"error": "Field 'question' is required."}),
            mimetype="application/json",
            status_code=400,
        )

    # 1. Embed the question
    query_vector = openai_client.get_embedding(question)

    # 2. Retrieve top-k relevant chunks
    hits = search_client.vector_search(query_vector, top_k=5)

    # 3. Build context string
    context = "\n\n---\n\n".join(
        f"[{h['document_name']}]\n{h['content']}" for h in hits
    )

    # 4. Generate grounded answer
    user_message = f"Context:\n{context}\n\nQuestion: {question}"
    answer = openai_client.chat_completion(SYSTEM_PROMPT, user_message)

    sources = [{"document": h["document_name"], "chunk": h["content"][:300]} for h in hits]

    return func.HttpResponse(
        json.dumps({"answer": answer, "sources": sources}),
        mimetype="application/json",
        status_code=200,
    )


# ---------------------------------------------------------------------------
# List documents
# ---------------------------------------------------------------------------

@app.route(route="documents", methods=["GET"])
def list_documents(req: func.HttpRequest) -> func.HttpResponse:
    docs = cosmos_client.list_documents()
    return func.HttpResponse(
        json.dumps({"documents": docs}),
        mimetype="application/json",
        status_code=200,
    )


# ---------------------------------------------------------------------------
# Excel Custom Skill — called by AI Search Indexer for .xlsx / .xls files
# ---------------------------------------------------------------------------

@app.route(route="excel_skill", methods=["POST"], auth_level=func.AuthLevel.ANONYMOUS)
def excel_skill(req: func.HttpRequest) -> func.HttpResponse:
    """AI Search Custom Skill: parse Excel file into natural-language row records."""
    body    = req.get_json()
    results = []

    for record in (body or {}).get("values", []):
        try:
            data             = record.get("data", {})
            source_blob_path = data.get("source_blob_path", "")
            source_file_name = data.get("source_file_name", "")

            # Strip container prefix from blob path if present
            blob_name = _strip_container_prefix(source_blob_path)
            raw_bytes = blob_client.download_document_by_path(blob_name)
            rows      = _parse_excel(raw_bytes, source_blob_path, source_file_name)

            results.append({
                "recordId": record["recordId"],
                "data":     {"excel_rows": rows},
                "errors":   [],
                "warnings": [],
            })
        except Exception as exc:
            logger.exception("excel_skill failed for record %s", record.get("recordId"))
            results.append({
                "recordId": record["recordId"],
                "data":     {"excel_rows": []},
                "errors":   [{"message": str(exc)}],
                "warnings": [],
            })

    return func.HttpResponse(json.dumps({"values": results}), mimetype="application/json", status_code=200)


def _strip_container_prefix(blob_path: str) -> str:
    """Remove 'https://...blob.core.windows.net/container/' prefix if present."""
    import re
    container = os.environ.get("AZURE_STORAGE_CONTAINER_NAME", "")
    # Remove URL prefix
    blob_path = re.sub(r"^https?://[^/]+/[^/]+/", "", blob_path)
    # Remove bare container prefix
    if blob_path.startswith(container + "/"):
        blob_path = blob_path[len(container) + 1:]
    return blob_path


def _parse_excel(raw_bytes: bytes, blob_path: str, file_name: str) -> list[dict]:
    import openpyxl
    wb   = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
    rows = []
    for sheet in wb.worksheets:
        _unmerge_cells(sheet)
        data = [r for r in sheet.iter_rows(values_only=True) if any(v is not None for v in r)]
        if len(data) < 2:
            continue
        headers, data_start = _detect_headers(data)
        for row in data[data_start:]:
            sentence = _row_to_sentence(headers, row, sheet.title)
            if not sentence:
                continue
            rows.append({
                "content":          sentence,
                "source_blob_path": blob_path,
                "source_file_name": file_name,
                "sheet_name":       sheet.title,
                "doc_type":         "excel",
                "product_name":     _extract_product_name(sentence),
            })
    return rows


def _unmerge_cells(sheet):
    for rng in list(sheet.merged_cells.ranges):
        value = sheet.cell(rng.min_row, rng.min_col).value
        sheet.unmerge_cells(str(rng))
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                sheet.cell(r, c).value = value


def _detect_headers(data: list) -> tuple[list[str], int]:
    def all_strings(row):
        return all(v is None or isinstance(v, str) for v in row) and any(v for v in row)

    if len(data) >= 2 and all_strings(data[0]) and all_strings(data[1]):
        headers = []
        for h1, h2 in zip(data[0], data[1]):
            h1 = str(h1).strip() if h1 else ""
            h2 = str(h2).strip() if h2 else ""
            if h1 and h2 and h1 != h2:
                headers.append(f"{h1} - {h2}")
            else:
                headers.append(h1 or h2)
        return headers, 2

    if all_strings(data[0]):
        return [str(h).strip() if h else "" for h in data[0]], 1

    from openpyxl.utils import get_column_letter
    return [get_column_letter(i + 1) for i in range(len(data[0]))], 0


def _row_to_sentence(headers: list[str], row: tuple, sheet_name: str) -> str:
    parts = []
    for h, v in zip(headers, row):
        if v is None or str(v).strip() == "":
            continue
        h = h.strip()
        v_str = str(v).strip()
        parts.append(f"{h}: {v_str}" if h else v_str)
    if not parts:
        return ""
    return f"[{sheet_name}] " + ", ".join(parts) + "."


def _extract_product_name(text: str) -> str:
    import re
    m = re.search(r"(?i)([\w\s]*(medical|health|insurance|plan|policy|coverage)[\w\s]*)", text)
    return m.group(1).strip()[:100] if m else ""


# ---------------------------------------------------------------------------
# Data Cleaning Custom Skill — called by both Excel and Document skillsets
# ---------------------------------------------------------------------------

@app.route(route="clean_document", methods=["POST"], auth_level=func.AuthLevel.ANONYMOUS)
def clean_document(req: func.HttpRequest) -> func.HttpResponse:
    """AI Search Custom Skill: clean document text based on doc_type."""
    body    = req.get_json()
    results = []

    for record in (body or {}).get("values", []):
        try:
            data     = record.get("data", {})
            content  = data.get("content", "") or ""
            doc_type = (data.get("doc_type", "") or "").lower()

            if "excel" in doc_type:
                cleaned, notes = _clean_excel_row(content)
            else:
                cleaned, notes = _clean_document_text(content)

            results.append({
                "recordId": record["recordId"],
                "data": {
                    "cleaned_content": cleaned,
                    "cleaning_notes":  ", ".join(notes),
                },
                "errors":   [],
                "warnings": [],
            })
        except Exception as exc:
            results.append({
                "recordId": record["recordId"],
                "data": {"cleaned_content": content, "cleaning_notes": f"error: {exc}"},
                "errors":   [{"message": str(exc)}],
                "warnings": [],
            })

    return func.HttpResponse(json.dumps({"values": results}), mimetype="application/json", status_code=200)


def _clean_document_text(text: str) -> tuple[str, list[str]]:
    import re
    notes = []

    # Normalize line endings
    text = text.replace("\r\n", "\n").replace("\r", "\n")

    # Remove control characters
    before = len(text)
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]", "", text)
    if len(text) < before:
        notes.append("removed_control_chars")

    # Remove watermark patterns
    wm_patterns = [r"(?im)^\s*CONFIDENTIAL\s*$", r"(?im)^\s*DRAFT\s*$", r"(?im)^\s*WATERMARK\s*$"]
    for p in wm_patterns:
        if re.search(p, text):
            text = re.sub(p, "", text)
            notes.append("removed_watermark")

    # Remove headers/footers: repeated short lines + page number patterns
    text = _remove_doc_headers_footers(text, notes)

    # Remove table of contents pages (lines ending with page numbers)
    toc_lines = [l for l in text.split("\n") if re.match(r".{5,}\s+\.{3,}\s*\d+\s*$", l)]
    if len(toc_lines) > 5:
        text = re.sub(r"(?m)^.{5,}\s+\.{3,}\s*\d+\s*$", "", text)
        notes.append("removed_toc")

    # Re-associate clause numbers with content
    # "Article 3\nThe insurer..." → "Article 3 The insurer..."
    text = re.sub(r"(?m)^((?:Article|Section|Clause|第)\s*\d+[\.\d]*)\s*\n+", r"\1 ", text)
    notes.append("reassociated_clauses")

    # Normalize table whitespace (multiple spaces → single space per line)
    text = re.sub(r"[ \t]{2,}", " ", text)
    # Collapse excessive blank lines
    text = re.sub(r"\n{3,}", "\n\n", text)

    return text.strip(), notes


def _remove_doc_headers_footers(text: str, notes: list[str]) -> str:
    import re
    from collections import Counter

    _NOISE = [
        r"(?i)^\s*page\s+\d+\s*(of\s+\d+)?\s*$",
        r"^\s*-\s*\d+\s*-\s*$",
        r"^\s*\d+\s*$",
        r"(?i)^\s*(all rights reserved|copyright\s+\d{4})\s*$",
    ]

    lines = text.split("\n")
    lines = [l for l in lines if not any(re.match(p, l) for p in _NOISE)]

    # Heuristic: lines appearing 3+ times are likely headers/footers
    counts   = Counter(l.strip() for l in lines if l.strip())
    repeated = {s for s, n in counts.items() if n >= 3 and len(s) < 200}
    if repeated:
        lines = [l for l in lines if l.strip() not in repeated]
        notes.append("removed_repeated_header_footer")

    return "\n".join(lines)


def _clean_excel_row(content: str) -> tuple[str, list[str]]:
    import re
    notes  = []
    if not content or not content.strip():
        return "", ["empty_record"]

    # Strip whitespace
    content = content.strip()

    # Normalize numeric values: remove currency symbols, standardize units
    content = re.sub(r"[¥$€£]\s*", "", content)
    content = re.sub(r"(\d)\s*,\s*(\d{3})", r"\1\2", content)  # 50,000 → 50000
    notes.append("normalized_numerics")

    # Remove formatting-only content
    if re.match(r"^[\s\-_=|/\\]+$", content):
        return "", ["formatting_only"]

    return content.strip(), notes


# ---------------------------------------------------------------------------
# Agent Query — orchestrates retrieval + generation with tool calling
# ---------------------------------------------------------------------------

_AGENT_SYSTEM_PROMPT = """You are an expert insurance advisor with access to a health and medical insurance knowledge base.
Answer questions accurately based on policy documents.

Rules:
1. Always use the search_knowledge_base tool to find relevant policy information before answering.
2. If asked about a specific document, use generate_sas_url to provide a download link.
3. Be precise with coverage amounts, percentages, deductibles, and policy terms.
4. If information is not found in the knowledge base, clearly state that.
5. Always end your response with a Reference section listing all sources used.

Response format (always end with this):
---
Reference: {source_file_name} (Page {page_number} or Sheet: {sheet_name})
Link: {sas_url} (valid for 24 hours)
"""

_AGENT_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "search_knowledge_base",
            "description": "Search the insurance knowledge base using hybrid search (vector + keyword + semantic reranker). Returns relevant policy content with source metadata.",
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "The search query, e.g. 'inpatient surgery coverage limit Basic Plan'"
                    },
                    "doc_type": {
                        "type": "string",
                        "description": "Optional filter: 'pdf', 'excel', 'word'. Leave empty to search all.",
                        "enum": ["pdf", "excel", "word", ""]
                    },
                    "product_name": {
                        "type": "string",
                        "description": "Optional: filter by insurance product name."
                    }
                },
                "required": ["query"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "generate_sas_url",
            "description": "Generate a time-limited SAS URL for a Blob Storage file so the user can download the source document.",
            "parameters": {
                "type": "object",
                "properties": {
                    "blob_path": {
                        "type": "string",
                        "description": "The blob path or URL from source_blob_path field in search results."
                    },
                    "expiry_hours": {
                        "type": "integer",
                        "description": "URL validity in hours. Default 24.",
                        "default": 24
                    }
                },
                "required": ["blob_path"]
            }
        }
    }
]


@app.route(route="agent_query", methods=["POST"])
def agent_query(req: func.HttpRequest) -> func.HttpResponse:
    """Agentic RAG endpoint: LLM decides when to search + which SAS URLs to generate."""
    body     = req.get_json()
    question = (body or {}).get("question", "").strip()
    if not question:
        return func.HttpResponse(
            json.dumps({"error": "Field 'question' is required."}),
            mimetype="application/json",
            status_code=400,
        )

    messages = [
        {"role": "system", "content": _AGENT_SYSTEM_PROMPT},
        {"role": "user",   "content": question},
    ]

    max_iterations = 6
    for _ in range(max_iterations):
        choice = openai_client.chat_with_tools(messages, _AGENT_TOOLS)

        if choice.finish_reason == "tool_calls":
            messages.append(choice.message.model_dump())
            for tc in choice.message.tool_calls:
                result = _execute_agent_tool(tc)
                messages.append({
                    "role":         "tool",
                    "tool_call_id": tc.id,
                    "content":      json.dumps(result),
                })
        else:
            answer = choice.message.content or ""
            return func.HttpResponse(
                json.dumps({"answer": answer}),
                mimetype="application/json",
                status_code=200,
            )

    return func.HttpResponse(
        json.dumps({"error": "Agent did not converge within iteration limit."}),
        mimetype="application/json",
        status_code=500,
    )


def _execute_agent_tool(tool_call) -> dict:
    import json as _json
    name = tool_call.function.name
    args = _json.loads(tool_call.function.arguments or "{}")

    if name == "search_knowledge_base":
        query        = args.get("query", "")
        doc_type     = args.get("doc_type", "")
        product_name = args.get("product_name", "")

        query_vector = openai_client.get_embedding(query)

        filter_parts = []
        if doc_type:
            filter_parts.append(f"doc_type eq '{doc_type}'")
        if product_name:
            filter_parts.append(f"product_name eq '{product_name}'")
        filter_expr = " and ".join(filter_parts) if filter_parts else None

        hits = search_client.hybrid_search(query, query_vector, top_k=5, filter_expr=filter_expr)
        return {"results": hits}

    if name == "generate_sas_url":
        raw_path     = args.get("blob_path", "")
        expiry_hours = int(args.get("expiry_hours", 24))
        blob_name    = _strip_container_prefix(raw_path)
        sas_url      = blob_client.generate_sas_url(blob_name, expiry_hours)
        return {"sas_url": sas_url, "blob_path": blob_name}

    return {"error": f"Unknown tool: {name}"}


# ---------------------------------------------------------------------------
# Setup Indexer — one-time admin call to create/update the full pipeline
# ---------------------------------------------------------------------------

@app.route(route="setup_indexer", methods=["POST"])
def setup_indexer(req: func.HttpRequest) -> func.HttpResponse:
    from shared.indexer_setup import setup_indexer_pipeline
    try:
        search_client.ensure_index()
        setup_indexer_pipeline()
        return func.HttpResponse(
            json.dumps({"status": "Index and indexer pipeline created and started"}),
            mimetype="application/json",
            status_code=200,
        )
    except Exception as exc:
        logger.exception("Failed to setup indexer pipeline")
        return func.HttpResponse(
            json.dumps({"error": str(exc)}),
            mimetype="application/json",
            status_code=500,
        )


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _extract_text(file_name: str, data: bytes) -> str:
    if file_name.lower().endswith(".pdf"):
        import PyPDF2
        reader = PyPDF2.PdfReader(io.BytesIO(data))
        return "\n".join(page.extract_text() or "" for page in reader.pages)
    # Plain text / fallback
    return data.decode("utf-8", errors="replace")


def _chunk_text(text: str, chunk_size: int = 500, overlap: int = 50) -> list[str]:
    """Split text into overlapping token-aware chunks."""
    import tiktoken
    enc = tiktoken.get_encoding("cl100k_base")
    tokens = enc.encode(text)
    chunks = []
    start = 0
    while start < len(tokens):
        end = min(start + chunk_size, len(tokens))
        chunk_tokens = tokens[start:end]
        chunks.append(enc.decode(chunk_tokens))
        start += chunk_size - overlap
    return [c for c in chunks if c.strip()]
