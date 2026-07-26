import io
import json
import logging
import os
import uuid

from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__), "..", ".env"))

from fastapi import FastAPI, File, UploadFile, HTTPException, BackgroundTasks, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from pydantic import BaseModel
from typing import List

from shared import blob_client, search_client, openai_client, cosmos_client, user_client
import langfuse_prompts
import langfuse_tracing

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(title="LayaKB API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("shutdown")
def _on_shutdown():
    langfuse_tracing.flush_langfuse()


def _system_prompt() -> str:
    """Current SYSTEM_PROMPT text, fetched from Langfuse Prompt Management (TTL-cached, falls back to a hardcoded copy)."""
    return langfuse_prompts.get_prompt(langfuse_prompts.SYSTEM_PROMPT_NAME)


def _agent_system_prompt() -> str:
    """Current agent SYSTEM_PROMPT text, fetched from Langfuse Prompt Management (TTL-cached, falls back to a hardcoded copy)."""
    return langfuse_prompts.get_prompt(langfuse_prompts.AGENT_SYSTEM_PROMPT_NAME)

_AGENT_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "search_knowledge_base",
            "description": "Search the insurance knowledge base using hybrid search (vector + keyword + semantic reranker).",
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {"type": "string"},
                    "doc_type": {"type": "string", "enum": ["pdf", "excel", "word", ""]},
                    "product_name": {"type": "string"},
                },
                "required": ["query"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "generate_sas_url",
            "description": "Generate a time-limited SAS URL for a Blob Storage file.",
            "parameters": {
                "type": "object",
                "properties": {
                    "blob_path": {"type": "string"},
                    "expiry_hours": {"type": "integer", "default": 24},
                },
                "required": ["blob_path"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "get_user_profile",
            "description": "Get a customer's full profile including personal details, all policies, and claims history.",
            "parameters": {
                "type": "object",
                "properties": {
                    "user_id": {"type": "string", "description": "Customer ID, e.g. USR001"},
                },
                "required": ["user_id"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "get_user_policies",
            "description": "Get all policies (active and historical) for a customer.",
            "parameters": {
                "type": "object",
                "properties": {
                    "user_id": {"type": "string"},
                },
                "required": ["user_id"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "get_user_claims",
            "description": "Get the full claims history for a customer.",
            "parameters": {
                "type": "object",
                "properties": {
                    "user_id": {"type": "string"},
                },
                "required": ["user_id"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "search_users",
            "description": "Search customers by name, email, or policy number prefix.",
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {"type": "string"},
                },
                "required": ["query"],
            },
        },
    },
]


# ---------------------------------------------------------------------------
# Health check
# ---------------------------------------------------------------------------

@app.get("/api/health")
def health():
    return {"status": "ok"}


# ---------------------------------------------------------------------------
# Ingest helpers
# ---------------------------------------------------------------------------

def _index_document(doc_id: str, file_name: str, raw_bytes: bytes):
    """Background task: embed and index a document already uploaded to blob."""
    try:
        text = _extract_text(file_name, raw_bytes)
        chunks = _chunk_text(text)

        search_client.ensure_index()

        search_docs = []
        for i, chunk in enumerate(chunks):
            vector = openai_client.get_embedding(chunk)
            search_docs.append({
                "id": f"{doc_id}_{i}",
                "document_id": doc_id,
                "source_file_name": file_name,
                "content": chunk,
                "content_vector": vector,
            })
        search_client.upsert_chunks(search_docs)

        cosmos_client.update_document_status(doc_id, status="indexed", chunks=len(chunks))
        logger.info("Indexed document %s (%d chunks)", file_name, len(chunks))
    except Exception:
        logger.exception("Failed to index document %s", file_name)
        cosmos_client.update_document_status(doc_id, status="failed")


# ---------------------------------------------------------------------------
# Ingest a single document (async)
# ---------------------------------------------------------------------------

@app.post("/api/ingest")
async def ingest(background_tasks: BackgroundTasks, file: UploadFile = File(...)):
    doc_id = str(uuid.uuid4())
    file_name = file.filename or f"{doc_id}.bin"
    raw_bytes = await file.read()

    blob_client.upload_document(file_name, raw_bytes)

    metadata = {
        "id": doc_id,
        "name": file_name,
        "size": len(raw_bytes),
        "chunks": 0,
        "status": "processing",
    }
    cosmos_client.save_document_metadata(metadata)

    background_tasks.add_task(_index_document, doc_id, file_name, raw_bytes)

    logger.info("Accepted document %s for async indexing", file_name)
    return JSONResponse(content=metadata)


# ---------------------------------------------------------------------------
# Batch ingest (async)
# ---------------------------------------------------------------------------

@app.post("/api/ingest/batch")
async def ingest_batch(background_tasks: BackgroundTasks, files: List[UploadFile] = File(...)):
    results = []
    for file in files:
        doc_id = str(uuid.uuid4())
        file_name = file.filename or f"{doc_id}.bin"
        raw_bytes = await file.read()

        blob_client.upload_document(file_name, raw_bytes)

        metadata = {
            "id": doc_id,
            "name": file_name,
            "size": len(raw_bytes),
            "chunks": 0,
            "status": "processing",
        }
        cosmos_client.save_document_metadata(metadata)
        background_tasks.add_task(_index_document, doc_id, file_name, raw_bytes)
        results.append(metadata)
        logger.info("Accepted document %s for async indexing", file_name)

    return JSONResponse(content={"documents": results})


# ---------------------------------------------------------------------------
# Query the knowledge base
# ---------------------------------------------------------------------------

class QueryRequest(BaseModel):
    question: str
    user_id: str | None = None


@app.post("/api/query")
def query(request: Request, body: QueryRequest):
    question = body.question.strip()
    if not question:
        raise HTTPException(status_code=400, detail="Field 'question' is required.")

    trace_metadata = langfuse_tracing.extract_trace_metadata(request.headers)
    with langfuse_tracing.observe_query(question, body.user_id, trace_metadata) as span:
        query_vector = openai_client.get_embedding(question)
        with langfuse_tracing.trace_tool_call("search_knowledge_base", {"query": question, "top_k": 5}):
            hits = search_client.vector_search(query_vector, top_k=5)

        kb_context = "\n\n---\n\n".join(
            f"[{h['source_file_name']}]\n{h['content']}" for h in hits
        )

        user_context = ""
        if body.user_id:
            user_context = user_client.build_user_context(body.user_id)

        if user_context:
            user_message = (
                f"Customer context:\n{user_context}\n\n"
                f"Knowledge base context:\n{kb_context}\n\n"
                f"Question: {question}"
            )
        else:
            user_message = f"Context:\n{kb_context}\n\nQuestion: {question}"

        model = os.environ.get("ARK_CHAT_MODEL", "")
        with langfuse_tracing.trace_llm_call(model, message_count=2):
            answer = openai_client.chat_completion(_system_prompt(), user_message)

        sources = [{"document": h["source_file_name"], "chunk": h["content"][:300]} for h in hits]
        if span:
            span.update(output={"answer": answer, "sources": sources})
        return {"answer": answer, "sources": sources}


@app.get("/api/users")
def list_users():
    return {"users": user_client.list_users()}


@app.get("/api/users/{user_id}")
def get_user(user_id: str):
    user = user_client.get_user_by_id(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user


# ---------------------------------------------------------------------------
# List documents
# ---------------------------------------------------------------------------

@app.get("/api/documents")
def list_documents():
    docs = cosmos_client.list_documents()
    return {"documents": docs}


# ---------------------------------------------------------------------------
# Delete a document
# ---------------------------------------------------------------------------

@app.delete("/api/documents/{doc_id}")
def delete_document(doc_id: str):
    meta = cosmos_client.get_document_metadata(doc_id)
    if not meta:
        raise HTTPException(status_code=404, detail="Document not found")

    errors = []

    try:
        search_client.delete_document_chunks(doc_id)
    except Exception as e:
        logger.exception("Failed to delete search chunks for %s", doc_id)
        errors.append(f"search: {e}")

    try:
        blob_client.delete_document(meta["name"])
    except Exception as e:
        logger.exception("Failed to delete blob for %s", doc_id)
        errors.append(f"blob: {e}")

    try:
        cosmos_client.delete_document_metadata(doc_id)
    except Exception as e:
        logger.exception("Failed to delete metadata for %s", doc_id)
        errors.append(f"cosmos: {e}")

    if errors:
        raise HTTPException(status_code=500, detail={"errors": errors})

    return {"deleted": doc_id}


# ---------------------------------------------------------------------------
# Excel Custom Skill — called by AI Search Indexer
# ---------------------------------------------------------------------------

@app.post("/api/excel_skill")
async def excel_skill(request_body: dict):
    results = []
    for record in request_body.get("values", []):
        try:
            data = record.get("data", {})
            source_blob_path = data.get("source_blob_path", "")
            source_file_name = data.get("source_file_name", "")
            blob_name = _strip_container_prefix(source_blob_path)
            raw_bytes = blob_client.download_document_by_path(blob_name)
            rows = _parse_excel(raw_bytes, source_blob_path, source_file_name)
            results.append({
                "recordId": record["recordId"],
                "data": {"excel_rows": rows},
                "errors": [],
                "warnings": [],
            })
        except Exception as exc:
            logger.exception("excel_skill failed for record %s", record.get("recordId"))
            results.append({
                "recordId": record["recordId"],
                "data": {"excel_rows": []},
                "errors": [{"message": str(exc)}],
                "warnings": [],
            })
    return {"values": results}


# ---------------------------------------------------------------------------
# Data Cleaning Custom Skill — called by AI Search Indexer
# ---------------------------------------------------------------------------

@app.post("/api/clean_document")
async def clean_document(request_body: dict):
    results = []
    for record in request_body.get("values", []):
        try:
            data = record.get("data", {})
            content = data.get("content", "") or ""
            doc_type = (data.get("doc_type", "") or "").lower()

            if "excel" in doc_type:
                cleaned, notes = _clean_excel_row(content)
            else:
                cleaned, notes = _clean_document_text(content)

            results.append({
                "recordId": record["recordId"],
                "data": {
                    "cleaned_content": cleaned,
                    "cleaning_notes": ", ".join(notes),
                },
                "errors": [],
                "warnings": [],
            })
        except Exception as exc:
            results.append({
                "recordId": record["recordId"],
                "data": {"cleaned_content": content, "cleaning_notes": f"error: {exc}"},
                "errors": [{"message": str(exc)}],
                "warnings": [],
            })
    return {"values": results}


# ---------------------------------------------------------------------------
# Agent Query
# ---------------------------------------------------------------------------

@app.post("/api/agent_query")
def agent_query(request: Request, body: QueryRequest):
    question = body.question.strip()
    if not question:
        raise HTTPException(status_code=400, detail="Field 'question' is required.")

    trace_metadata = langfuse_tracing.extract_trace_metadata(request.headers)
    with langfuse_tracing.observe_agent_query(question, body.user_id, trace_metadata) as span:
        # Inject customer context if a user_id was provided
        user_context = ""
        if body.user_id:
            user_context = user_client.build_user_context(body.user_id)

        user_message = question
        if user_context:
            user_message = f"Customer context:\n{user_context}\n\nQuestion: {question}"

        messages = [
            {"role": "system", "content": _agent_system_prompt()},
            {"role": "user", "content": user_message},
        ]

        model = os.environ.get("ARK_CHAT_MODEL", "")
        for iteration in range(6):
            with langfuse_tracing.trace_agent_loop_iteration(iteration, len(messages)):
                with langfuse_tracing.trace_llm_call(model, message_count=len(messages)):
                    choice = openai_client.chat_with_tools(messages, _AGENT_TOOLS)

                if choice.finish_reason == "tool_calls":
                    messages.append(choice.message.model_dump())
                    for tc in choice.message.tool_calls:
                        with langfuse_tracing.trace_tool_call(
                            tc.function.name, json.loads(tc.function.arguments or "{}")
                        ):
                            result = _execute_agent_tool(tc)
                        messages.append({
                            "role": "tool",
                            "tool_call_id": tc.id,
                            "content": json.dumps(result),
                        })
                else:
                    answer = choice.message.content or ""
                    if span:
                        span.update(output={"answer": answer})
                    return {"answer": answer}

        raise HTTPException(status_code=500, detail="Agent did not converge within iteration limit.")


# ---------------------------------------------------------------------------
# Evaluation
# ---------------------------------------------------------------------------

class EvalRequest(BaseModel):
    category: str | None = None
    ids: list[str] | None = None


@app.post("/api/evaluate")
def run_evaluation(body: EvalRequest = EvalRequest()):
    """
    Run LLM-as-judge evaluation against the test suite.
    Optionally filter by category or specific test IDs.
    Long-running — may take several minutes for the full suite.
    """
    from evaluation.test_cases import TEST_CASES
    from evaluation.evaluator import evaluate_case
    import importlib

    main_mod = importlib.import_module("main")

    def _agent_fn(question: str, uid: str | None) -> dict:
        ctx = user_client.build_user_context(uid) if uid else ""
        content = f"Customer context:\n{ctx}\n\nQuestion: {question}" if ctx else question
        messages = [
            {"role": "system", "content": main_mod._agent_system_prompt()},
            {"role": "user",   "content": content},
        ]
        for _ in range(6):
            choice = openai_client.chat_with_tools(messages, main_mod._AGENT_TOOLS)
            if choice.finish_reason == "tool_calls":
                messages.append(choice.message.model_dump())
                for tc in choice.message.tool_calls:
                    result = _execute_agent_tool(tc)
                    messages.append({"role": "tool", "tool_call_id": tc.id, "content": json.dumps(result)})
            else:
                return {"answer": choice.message.content or "", "sources": []}
        return {"answer": "Agent did not converge.", "sources": []}

    cases = TEST_CASES
    if body.category:
        cases = [c for c in cases if c["category"] == body.category]
    if body.ids:
        cases = [c for c in cases if c["id"] in body.ids]

    results = []
    for case in cases:
        results.append(evaluate_case(case, _agent_fn))

    total = len(results)
    overall_avg = round(sum(r["overall"] for r in results) / total, 3) if total else 0
    hard_passed = sum(1 for r in results if r["hard_pass"])

    by_category: dict[str, dict] = {}
    for r in results:
        cat = r["category"]
        if cat not in by_category:
            by_category[cat] = {"scores": [], "count": 0}
        by_category[cat]["scores"].append(r["overall"])
        by_category[cat]["count"] += 1
    category_summary = {
        cat: {"avg": round(sum(v["scores"]) / v["count"], 3), "count": v["count"]}
        for cat, v in by_category.items()
    }

    return {
        "summary": {
            "total": total,
            "hard_passed": hard_passed,
            "overall_avg": overall_avg,
            "by_category": category_summary,
        },
        "results": results,
    }


# ---------------------------------------------------------------------------
# Setup Indexer
# ---------------------------------------------------------------------------

@app.post("/api/setup_indexer")
def setup_indexer():
    from shared.indexer_setup import setup_indexer_pipeline
    try:
        search_client.ensure_index()
        setup_indexer_pipeline()
        return {"status": "Index and indexer pipeline created and started"}
    except Exception as exc:
        logger.exception("Failed to setup indexer pipeline")
        raise HTTPException(status_code=500, detail=str(exc))


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _extract_text(file_name: str, data: bytes) -> str:
    if file_name.lower().endswith(".pdf"):
        import PyPDF2
        reader = PyPDF2.PdfReader(io.BytesIO(data))
        return "\n".join(page.extract_text() or "" for page in reader.pages)
    return data.decode("utf-8", errors="replace")


def _chunk_text(text: str, chunk_size: int = 500, overlap: int = 50) -> list[str]:
    import tiktoken
    enc = tiktoken.get_encoding("cl100k_base")
    tokens = enc.encode(text)
    chunks = []
    start = 0
    while start < len(tokens):
        end = min(start + chunk_size, len(tokens))
        chunks.append(enc.decode(tokens[start:end]))
        start += chunk_size - overlap
    return [c for c in chunks if c.strip()]


def _strip_container_prefix(blob_path: str) -> str:
    import re
    container = os.environ.get("AZURE_STORAGE_CONTAINER_NAME", "")
    blob_path = re.sub(r"^https?://[^/]+/[^/]+/", "", blob_path)
    if blob_path.startswith(container + "/"):
        blob_path = blob_path[len(container) + 1:]
    return blob_path


def _parse_excel(raw_bytes: bytes, blob_path: str, file_name: str) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
    rows = []
    for sheet in wb.worksheets:
        _unmerge_cells(sheet)
        data = [r for r in sheet.iter_rows(values_only=True) if any(v is not None for v in r)]
        if len(data) < 2:
            continue
        headers, data_start = _detect_headers(data)
        for row in data[data_start:]:
            sentence = _row_to_sentence(headers, row, sheet.title)
            if not sentence:
                continue
            rows.append({
                "content": sentence,
                "source_blob_path": blob_path,
                "source_file_name": file_name,
                "sheet_name": sheet.title,
                "doc_type": "excel",
                "product_name": _extract_product_name(sentence),
            })
    return rows


def _unmerge_cells(sheet):
    for rng in list(sheet.merged_cells.ranges):
        value = sheet.cell(rng.min_row, rng.min_col).value
        sheet.unmerge_cells(str(rng))
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                sheet.cell(r, c).value = value


def _detect_headers(data: list) -> tuple[list[str], int]:
    def all_strings(row):
        return all(v is None or isinstance(v, str) for v in row) and any(v for v in row)

    if len(data) >= 2 and all_strings(data[0]) and all_strings(data[1]):
        headers = []
        for h1, h2 in zip(data[0], data[1]):
            h1 = str(h1).strip() if h1 else ""
            h2 = str(h2).strip() if h2 else ""
            headers.append(f"{h1} - {h2}" if h1 and h2 and h1 != h2 else h1 or h2)
        return headers, 2

    if all_strings(data[0]):
        return [str(h).strip() if h else "" for h in data[0]], 1

    from openpyxl.utils import get_column_letter
    return [get_column_letter(i + 1) for i in range(len(data[0]))], 0


def _row_to_sentence(headers: list[str], row: tuple, sheet_name: str) -> str:
    parts = []
    for h, v in zip(headers, row):
        if v is None or str(v).strip() == "":
            continue
        parts.append(f"{h.strip()}: {str(v).strip()}" if h.strip() else str(v).strip())
    if not parts:
        return ""
    return f"[{sheet_name}] " + ", ".join(parts) + "."


def _extract_product_name(text: str) -> str:
    import re
    m = re.search(r"(?i)([\w\s]*(medical|health|insurance|plan|policy|coverage)[\w\s]*)", text)
    return m.group(1).strip()[:100] if m else ""


def _clean_document_text(text: str) -> tuple[str, list[str]]:
    import re
    notes = []
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    before = len(text)
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]", "", text)
    if len(text) < before:
        notes.append("removed_control_chars")
    for p in [r"(?im)^\s*CONFIDENTIAL\s*$", r"(?im)^\s*DRAFT\s*$", r"(?im)^\s*WATERMARK\s*$"]:
        if re.search(p, text):
            text = re.sub(p, "", text)
            notes.append("removed_watermark")
    text = _remove_doc_headers_footers(text, notes)
    toc_lines = [l for l in text.split("\n") if re.match(r".{5,}\s+\.{3,}\s*\d+\s*$", l)]
    if len(toc_lines) > 5:
        text = re.sub(r"(?m)^.{5,}\s+\.{3,}\s*\d+\s*$", "", text)
        notes.append("removed_toc")
    text = re.sub(r"(?m)^((?:Article|Section|Clause|第)\s*\d+[\.\d]*)\s*\n+", r"\1 ", text)
    notes.append("reassociated_clauses")
    text = re.sub(r"[ \t]{2,}", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip(), notes


def _remove_doc_headers_footers(text: str, notes: list[str]) -> str:
    import re
    from collections import Counter
    _NOISE = [
        r"(?i)^\s*page\s+\d+\s*(of\s+\d+)?\s*$",
        r"^\s*-\s*\d+\s*-\s*$",
        r"^\s*\d+\s*$",
        r"(?i)^\s*(all rights reserved|copyright\s+\d{4})\s*$",
    ]
    lines = text.split("\n")
    lines = [l for l in lines if not any(re.match(p, l) for p in _NOISE)]
    counts = Counter(l.strip() for l in lines if l.strip())
    repeated = {s for s, n in counts.items() if n >= 3 and len(s) < 200}
    if repeated:
        lines = [l for l in lines if l.strip() not in repeated]
        notes.append("removed_repeated_header_footer")
    return "\n".join(lines)


def _clean_excel_row(content: str) -> tuple[str, list[str]]:
    import re
    notes = []
    if not content or not content.strip():
        return "", ["empty_record"]
    content = content.strip()
    content = re.sub(r"[¥$€£]\s*", "", content)
    content = re.sub(r"(\d)\s*,\s*(\d{3})", r"\1\2", content)
    notes.append("normalized_numerics")
    if re.match(r"^[\s\-_=|/\\]+$", content):
        return "", ["formatting_only"]
    return content.strip(), notes


def _execute_agent_tool(tool_call) -> dict:
    name = tool_call.function.name
    args = json.loads(tool_call.function.arguments or "{}")

    if name == "search_knowledge_base":
        query = args.get("query", "")
        doc_type = args.get("doc_type", "")
        product_name = args.get("product_name", "")
        query_vector = openai_client.get_embedding(query)
        filter_parts = []
        if doc_type:
            filter_parts.append(f"doc_type eq '{doc_type}'")
        if product_name:
            filter_parts.append(f"product_name eq '{product_name}'")
        filter_expr = " and ".join(filter_parts) if filter_parts else None
        hits = search_client.hybrid_search(query, query_vector, top_k=5, filter_expr=filter_expr)
        return {"results": hits}

    if name == "generate_sas_url":
        raw_path = args.get("blob_path", "")
        expiry_hours = int(args.get("expiry_hours", 24))
        blob_name = _strip_container_prefix(raw_path)
        sas_url = blob_client.generate_sas_url(blob_name, expiry_hours)
        return {"sas_url": sas_url, "blob_path": blob_name}

    if name == "get_user_profile":
        uid = args.get("user_id", "")
        u = user_client.get_user_by_id(uid)
        return u if u else {"error": f"No customer found with ID '{uid}'"}

    if name == "get_user_policies":
        uid = args.get("user_id", "")
        policies = user_client.get_user_policies(uid)
        if policies is None:
            return {"error": f"No customer found with ID '{uid}'"}
        return {"policies": policies}

    if name == "get_user_claims":
        uid = args.get("user_id", "")
        claims = user_client.get_user_claims(uid)
        if claims is None:
            return {"error": f"No customer found with ID '{uid}'"}
        return {"claims": claims}

    if name == "search_users":
        results = user_client.search_users(args.get("query", ""), limit=10)
        return {"users": results}

    return {"error": f"Unknown tool: {name}"}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=False)
